import sys,os   #Demo5_12.py
from PySide6.QtWidgets import (QApplication,QWidget,QMenuBar,QVBoxLayout,QListView, 
    QStyledItemDelegate,QDoubleSpinBox,QComboBox,QFileDialog,QTableView,QSplitter)
from PySide6.QtGui import QStandardItemModel,QStandardItem,QIcon
from PySide6.QtCore import Qt
from openpyxl import load_workbook

class comboBoxDelegate(QStyledItemDelegate):  #代理子类
    def __init__(self,parent=None):
        super().__init__(parent)
    def createEditor(self, parent, option, index):  #创建代理控件
        comBox = QComboBox(parent)
        male = QIcon("d:\\python\\male.png")
        female = QIcon("d:\\python\\female.png")
        comBox.addItem(male, "男")
        comBox.addItem(female, "女")
        comBox.setEditable(False)
        return comBox  #返回代理控件
    def setEditorData(self,comBox,index):  #设置代理控件的数据
        model=index.model()   #获取模型
        if model.data(index,Qt.DisplayRole) == "男":
            comBox.setCurrentIndex(0)
        else:
            comBox.setCurrentIndex(1)
    def setModelData(self,editor,model,index):  #把代理控件的数据写入数据模型中
        comboBox_index=editor.currentIndex()
        text=editor.itemText(comboBox_index)
        icon=editor.itemIcon(comboBox_index)
        model.setData(index,text,Qt.DisplayRole)
        model.setData(index,icon,Qt.DecorationRole)
    def updateEditorGeometry(self,editor, option, index): #设置代理控件的位置
        editor.setGeometry(option.rect) 
class doubleSpinBoxDelegate(QStyledItemDelegate):  #代理子类
    def __init__(self,parent=None):
        super().__init__(parent)
    def createEditor(self, parent, option, index):
        editor=QDoubleSpinBox(parent)
        editor.setDecimals(1)  #设置1位小数
        editor.setMinimum(0.0)
        editor.setMaximum(100.0)
        editor.setFrame(False)
        return editor
    def setEditorData(self,editor,index):
        model=index.model()
        text=model.data(index,Qt.DisplayRole)
        try:
            editor.setValue(float(text))
        except:
            editor.setValue(0.0)
    def setModelData(self,editor,model,index):
        value=editor.value()
        model.setData(index,str(value),Qt.DisplayRole)
    def updateEditorGeometry(self,editor, option, index):
        editor.setGeometry(option.rect)

class MyWindow(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent)
        self.standardModel = QStandardItemModel(self)
        self.widget_setupUi()
        self.male=QIcon("d:\\python\\male.png")
        self.female=QIcon("d:\\python\\female.png")
    def widget_setupUi(self):  #建立主程序界面
        menuBar = QMenuBar(self)  #菜单栏
        fileMenu = menuBar.addMenu("文件") #菜单
        self.action_open=fileMenu.addAction("打开")  #打开动作
        fileMenu.addSeparator()
        self.action_exit=fileMenu.addAction("退出")

        self.listView=QListView(self)  #列表视图控件
        self.listView.setMaximumWidth(100)
        self.tableView=QTableView(self)  #表格视图控件
        self.tableView.setAlternatingRowColors(True)  #交替颜色

        h_splitter=QSplitter(Qt.Horizontal)  #布局
        h_splitter.addWidget(self.listView)
        h_splitter.addWidget(self.tableView)
        v=QVBoxLayout(self)
        v.addWidget(menuBar,0)
        v.addWidget(h_splitter,1)

        self.action_open.triggered.connect(self.action_open_triggered)  #信号与槽连接
        self.action_exit.triggered.connect(self.close)  # 信号与槽连接
        self.listView.clicked.connect(self.listView_clicked)  #信号与槽连接
    def action_open_triggered(self):  #打开Excel文件的槽函数
        fileName,fil=QFileDialog.getOpenFileName(self,"打开文件","d:\\","Excel文件(*.xlsx)")
        if os.path.exists(fileName):
            self.standardModel.clear()
            wbook=load_workbook(fileName)
            for sheetname in wbook.sheetnames:
                wsheet=wbook[sheetname]
                cell_range = wsheet[wsheet.dimensions]  # 按行排列的单元格对象元组
                score = list()  #记录Excel工作表格中的数据
                for rowCells in cell_range:  # rowCells是excel行单元格元组
                    temp = list()  # 临时列表
                    for cell in rowCells:  # cell是单元格对象
                        temp.append(str(cell.value))
                    score.append(temp)
                parent_item = QStandardItem(sheetname)  #根索引下的顶层数据项
                parent_item.setColumnCount(len(score[0]))  #设置顶层数据项下列的数量
                for i in range(1,len(score)):
                    items_temp=list()  #临时列表
                    for j in score[i]:
                        child_item=QStandardItem(j)  #子数据项
                        child_item.setTextAlignment(Qt.AlignCenter)
                        if j=="男":
                            child_item.setIcon(self.male)  #设置图标
                        elif j=="女":
                            child_item.setIcon(self.female)
                        items_temp.append(child_item)
                    parent_item.appendRow(items_temp)  #将子数据列表添加到顶层数据项中
                self.standardModel.appendRow(parent_item)
            self.standardModel.setHorizontalHeaderLabels(score[0])  #设置水平表头
            self.listView.setModel(self.standardModel)  # 设置列表视图控件的数据模型
            self.tableView.setModel(self.standardModel)  #设置表格视图控件的数据模型
            index = self.standardModel.index(0, 0)
            self.listView_clicked(index)  #初始时刻表格视图控件和树视图控件指向的根索引
    def listView_clicked(self,index):  #列表视图控件的单击槽函数
        item=self.standardModel.itemFromIndex(index)
        if item.hasChildren():
            self.tableView.setRootIndex(index)
            row_count=item.rowCount()
            label=list()
            for i in range(1,row_count+1):  #不含表头
                label.append(str(i))
            self.standardModel.setVerticalHeaderLabels(label)  #设置列表头显示的文字

        comBoxDelegate=comboBoxDelegate(self)
        doubleSpinDelegate=doubleSpinBoxDelegate(self)
        header=self.tableView.horizontalHeader()
        for i in range(header.count()):  #根据表头的名称设置代理类型
            header_text=self.standardModel.horizontalHeaderItem(i).data(Qt.DisplayRole)
            if header_text in ["语文","数学","物理","化学","历史","地理"]:
                self.tableView.setItemDelegateForColumn(i, doubleSpinDelegate)  #设置代理
            elif header_text=="性别":
                self.tableView.setItemDelegateForColumn(i,comBoxDelegate)  #设置代理
if __name__ == '__main__':
    app=QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec())
