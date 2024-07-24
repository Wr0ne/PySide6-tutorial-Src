import sys,os   #Demo5_6.py
from PySide6.QtWidgets import (QApplication,QWidget,QMenuBar,QVBoxLayout,
                             QFileDialog,QTableWidget,QTableWidgetItem)
from openpyxl import load_workbook,Workbook

class MyWindow(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent)
        self.setupUi()
    def setupUi(self):  #建立主程序界面
        menuBar = QMenuBar(self)
        fileMenu = menuBar.addMenu("文件") #菜单
        self.action_open=fileMenu.addAction("打开")  #动作
        self.action_saveAs=fileMenu.addAction("另存")  #动作
        fileMenu.addSeparator()
        self.action_exit=fileMenu.addAction("退出")
        statisticMenu=menuBar.addMenu("统计") #菜单
        self.action_total=statisticMenu.addAction("插入总成绩")  #动作
        self.action_average = statisticMenu.addAction("插入平均分")  #动作
        self.action_saveAs.setEnabled(False)
        self.action_total.setEnabled(False)
        self.action_average.setEnabled(False)

        self.tableWidget=QTableWidget(self)  #表格控件
        v=QVBoxLayout(self)
        v.addWidget(menuBar)
        v.addWidget(self.tableWidget)

        self.action_open.triggered.connect(self.action_open_triggered)  #信号与槽连接
        self.action_saveAs.triggered.connect(self.action_saveAs_triggered)  #信号与槽连接
        self.action_exit.triggered.connect(self.close)  #信号与槽连接
        self.action_total.triggered.connect(self.action_total_triggered)  #信号与槽连接
        self.action_average.triggered.connect(self.action_average_triggered)  #信号与槽连接
    def action_open_triggered(self):  #打开Excel文件
        score=list()  #读取数据后，保存数据的列表
        fileName,fil=QFileDialog.getOpenFileName(self,"打开文件","d:\\","Excel文件(*.xlsx)")
        if os.path.exists(fileName):
            wbook=load_workbook(fileName)
            wsheet = wbook.active
            cell_range = wsheet[wsheet.dimensions]  # 按行排列的单元格对象元组

            for i in cell_range:  # i是Excel行单元格元组
                temp = list()  # 临时列表
                for j in i:  # j是单元格对象
                    temp.append(str(j.value))
                score.append(temp)
            row_count = len(score) - 1  # 行数，不包含表头
            column_count = len(score[0])  # 列数

            self.tableWidget.setRowCount(row_count)
            self.tableWidget.setColumnCount(column_count)
            self.tableWidget.setHorizontalHeaderLabels(score[0])
            for i in range(row_count):
                for j in range(column_count):
                    cell = QTableWidgetItem()
                    cell.setText(score[i + 1][j])
                    self.tableWidget.setItem(i, j, cell)
            self.action_saveAs.setEnabled(True)
            self.action_total.setEnabled(True)
            self.action_average.setEnabled(True)
    def action_saveAs_triggered(self):  #另存
        score = list()
        fileName,fil = QFileDialog.getSaveFileName(self,"保存文件","d:\\","Excel文件(*.xlsx)")
        if fileName!="":
            temp = list()
            for j in  range(self.tableWidget.columnCount()):
                temp.append(self.tableWidget.horizontalHeaderItem(j).text())
            score.append(temp)
            for i in range(self.tableWidget.rowCount()):
                temp=list()
                for j in range(self.tableWidget.columnCount()):
                    temp.append(self.tableWidget.item(i,j).text())
                score.append(temp)
            wbook=Workbook()
            wsheet=wbook.create_sheet("学生成绩",0)
            for i in score:
                wsheet.append(i)
            wbook.save(fileName)
    def action_total_triggered(self):  #计算总成绩
        column=self.tableWidget.columnCount()
        self.tableWidget.insertColumn(column)
        item= QTableWidgetItem("总成绩")
        self.tableWidget.setHorizontalHeaderItem(column,item)
        for i in range(self.tableWidget.rowCount()):
            total=0
            for j in range(2,6):
                total=total+int(self.tableWidget.item(i,j).text())
            item=QTableWidgetItem(str(total))
            self.tableWidget.setItem(i,column,item)
    def action_average_triggered(self):  #计算平均成绩
        column=self.tableWidget.columnCount()
        self.tableWidget.insertColumn(column)
        item= QTableWidgetItem("平均成绩")
        self.tableWidget.setHorizontalHeaderItem(column,item)
        for i in range(self.tableWidget.rowCount()):
            total=0
            for j in range(2,6):
                total=total+int(self.tableWidget.item(i,j).text())
            item=QTableWidgetItem(str(total/4))
            self.tableWidget.setItem(i,column,item)
if __name__ == '__main__':
    app=QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec())
