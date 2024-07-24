import openpyxl, math  #Demo5_4.py
mybook = openpyxl.Workbook()
mysheet = mybook.active
mysheet.title = "正弦和余弦值"
mysheet["A1"] = "角度值（度）"
mysheet["B1"] = "正弦值"
mysheet["C1"] = "余弦值"
for i in range(360):
    mysheet.cell(row=i+2, column=1, value=i)
    mysheet.cell(row=i+2, column=2, value=math.sin(i*math.pi/180))
    mysheet.cell(row=i+2, column=3, value=math.cos(i*math.pi/180))
mybook.save("d:\\sin_cos.xlsx")
