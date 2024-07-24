import sys,os   #Demo5_11.py
from PySide6.QtWidgets import (QApplication,QWidget,QMenuBar,QVBoxLayout,QTreeView,QListView,
                             QFileDialog,QTableView,QSplitter)
from PySide6.QtGui import QStandardItemModel,QStandardItem,QIcon
from PySide6.QtCore import Qt,QModelIndex
from openpyxl import load_workbook,Workbook

class MyWindow(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent)
        self.standardModel = QStandardItemModel(self)  #标注数据模型
        self.widget_setupUi()
        self.male=QIcon("d:\\python\\male.png")   #图标
        self.female=QIcon("d:\\python\\female.png")  #图标
    def widget_setupUi(self):  #建立主程序界面
        menuBar = QMenuBar(self)  #菜单栏
        fileMenu = menuBar.addMenu("文件") #菜单
        self.action_open=fileMenu.addAction("打开")  #打开动作
        self.action_saveAs=fileMenu.addAction("另存")  #另存动作
        fileMenu.addSeparator()
        self.action_exit=fileMenu.addAction("退出")
        self.action_saveAs.setEnabled(False)

        self.listView=QListView(self)  #列表视图控件
        self.listView.setMaximumWidth(100)
        self.tableView=QTableView(self)  #表格视图控件
        self.treeView=QTreeView(self)  #树视图控件
        self.tableView.setAlternatingRowColors(True)  #交替颜色
        self.treeView.setAlternatingRowColors(True)
        h_splitter=QSplitter(Qt.Horizontal)  #布局
        v_splitter=QSplitter(Qt.Vertical)
        h_splitter.addWidget(self.listView)
        h_splitter.addWidget(v_splitter)
        v_splitter.addWidget(self.tableView)
        v_splitter.addWidget(self.treeView)
        v=QVBoxLayout(self)
        v.addWidget(menuBar,0)
        v.addWidget(h_splitter,1)
        v.setSpacing(0)
        self.action_open.triggered.connect(self.action_open_triggered)  #信号与槽连接
        self.action_saveAs.triggered.connect(self.action_saveAs_triggered) #信号与槽连接
        self.action_exit.triggered.connect(self.close)  # 信号与槽连接
        self.listView.clicked.connect(self.listView_clicked)  #信号与槽连接
    def action_open_triggered(self):  #打开Excel文件的槽函数
        fileName,fil=QFileDialog.getOpenFileName(self,"打开文件","d:\\","Excel文件(*.xlsx)")
        if os.path.exists(fileName):
            self.standardModel.clear()
            wbook=load_workbook(fileName)
            for sheetname in wbook.sheetnames:
                wsheet=wbook[sheetname]
                cell_range = wsheet[wsheet.dimensions]  # 按行排列的单元格对象元组
                score = list()  #记录excel工作表格中的数据
                for rowCells in cell_range:  # rowCells是Excel行单元格元组
                    temp = list()  # 临时列表
                    for cell in rowCells:  # cell是单元格对象
                        temp.append(str(cell.value))
                    score.append(temp)
                parent_item = QStandardItem(sheetname)  #根索引下的顶层数据项
                parent_item.setColumnCount(len(score[0])) #设置顶层数据项下列的数量
                for i in range(1,len(score)):
                    items_temp=list()  #临时列表
                    for j in score[i]:
                        child_item=QStandardItem(j)  #子数据项
                        child_item.setTextAlignment(Qt.AlignCenter)
                        if j=="男":
                            child_item.setIcon(self.male)  #设置图标
                        elif j=="女":
                            child_item.setIcon(self.female)
                        items_temp.append(child_item)
                    parent_item.appendRow(items_temp) #将子数据列表添加到顶层项中
                self.standardModel.appendRow(parent_item)
            self.standardModel.setHorizontalHeaderLabels(score[0]) #设置水平表头
            self.listView.setModel(self.standardModel)  # 设置列表视图控件的数据模型
            self.tableView.setModel(self.standardModel)  #设置表格视图控件的数据模型
            self.treeView.setModel(self.standardModel)   #设置树视图控件的数据模型
            self.action_saveAs.setEnabled(True)
            index = self.standardModel.index(0, 0)
            self.listView_clicked(index) #初始时刻表格视图和树视图控件指向的根索引
    def listView_clicked(self,index):  #列表视图控件的单击槽函数
        item=self.standardModel.itemFromIndex(index)
        if item.hasChildren():
            self.tableView.setRootIndex(index)
            self.treeView.collapseAll()
            self.treeView.expand(index)
            row_count=item.rowCount()
            label=list()
            for i in range(1,row_count+1):
                label.append(str(i))
            self.standardModel.setVerticalHeaderLabels(label) #设置列表头显示的文字
    def action_saveAs_triggered(self):  # 另存动作的槽函数
        sheet_count=self.standardModel.rowCount(QModelIndex()) #根索引下数据项的数量
        wbook=Workbook()
        for i in range(sheet_count):
            parent_index= self.standardModel.index(i,0,QModelIndex()) #根索引下顶层索引
            parent_item=self.standardModel.itemFromIndex(parent_index) 
            if parent_item.hasChildren():
                sheet_name=self.standardModel.data(parent_index,Qt.DisplayRole)
                wsheet = wbook.create_sheet(sheet_name,i)

                row_count=self.standardModel.rowCount(parent_index)
                column_count = self.standardModel.columnCount(parent_index)
                horizontal_header = list()
                for column in range(column_count):  #获取水平表头的文本
                    header_name=self.standardModel.headerData(column,
                                        Qt.Horizontal,Qt.DisplayRole)
                    horizontal_header.append(header_name)
                wsheet.append(horizontal_header)  #在工作表格中添加表头
                for row in range(row_count): #获取每个班级的数据
                    score=list()
                    for column in range(column_count): #获取每行的数据
                        child_index=parent_index.child(row,column)  #获取子索引
                        score.append(child_index.data(Qt.DisplayRole))  #获取数据
                    wsheet.append(score)  #在工作表格中添加数据
        fileName, fil = QFileDialog.getSaveFileName(self, "保存文件", "d:\\", 
"Excel文件(*.xlsx)")
        if fileName != "":
            wbook.save(fileName)
if __name__ == '__main__':
    app=QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec())
