import sys,os   #Demo5_9.py
from PySide6.QtWidgets import (QApplication,QWidget,QListView,QHBoxLayout,
                             QLabel,QPushButton,QVBoxLayout)
from PySide6.QtCore import QStringListModel,QModelIndex,Qt
from openpyxl import load_workbook

class MyWindow(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent)
        self.fileName = "d:\\python\\学生ID.xlsx"
        self.reference_Model=QStringListModel(self)  #从Excel中读取数据后，存储数据的模型
        self.selection_Model=QStringListModel(self)  #选择数据后，存储选择数据的模型
        self.setup_Ui()  #建立界面
        self.data_import()  #从Excel中读取数据
        self.view_clicked()  #单击视图控件，判断按钮是否激活或失效
    def setup_Ui(self):  #建立界面
        label1 = QLabel("学生名单")
        self.listView_1=QListView()  #列表视图控件，显示Excel中的数据的控件
        v1=QVBoxLayout()
        v1.addWidget(label1)
        v1.addWidget(self.listView_1)

        label2=QLabel("三好学生")
        self.listView_2=QListView()  #列表视图控件，显示选中的数据
        self.btn_add = QPushButton("添加")
        self.btn_insert = QPushButton("插入")
        self.btn_delete = QPushButton("删除")
        h1=QHBoxLayout()
        h1.addWidget(self.btn_add)
        h1.addWidget(self.btn_insert)
        h1.addWidget(self.btn_delete)
        v2=QVBoxLayout()
        v2.addWidget(label2)
        v2.addWidget(self.listView_2)
        v2.addLayout(h1)
        h2=QHBoxLayout(self)
        h2.addLayout(v1)
        h2.addLayout(v2)

        self.listView_1.setModel(self.reference_Model)  #设置模型
        self.listView_2.setModel(self.selection_Model)  #设置模型
        self.listView_1.setSelectionMode(QListView.ExtendedSelection) #设置选择模式
        self.listView_2.setSelectionMode(QListView.ExtendedSelection) #设置选择模式

        self.btn_add.clicked.connect(self.btn_add_clicked)
        self.btn_insert.clicked.connect(self.btn_insert_clicked)
        self.btn_delete.clicked.connect(self.btn_delete_clicked)
        self.listView_1.clicked.connect(self.view_clicked)
        self.listView_2.clicked.connect(self.view_clicked)
    def data_import(self):
        if os.path.exists(self.fileName):
            wbook = load_workbook(self.fileName)
            if "ID" in wbook.sheetnames:
                wsheet = wbook["ID"]
                cell_range = wsheet[wsheet.dimensions]  #获取Excel中数据存储的范围
                student=list()
                for cell_row in cell_range:  # cell_row是Excel行单元格元组
                    string = ""
                    for cell in cell_row:
                        string=string+str(cell.value)+"  "  #获取excel单元格中的数据
                    student.append(string.strip())
                self.reference_Model.setStringList(student)  #在模型中添加数据列表
    def btn_add_clicked(self):  #添加按钮的槽函数
        while len(self.listView_1.selectedIndexes()):
            selectedIndexes=self.listView_1.selectedIndexes()
            index= selectedIndexes[0]
            string = self.reference_Model.data(index,Qt.DisplayRole)  #获取数据
            self.reference_Model.removeRow(index.row(),QModelIndex())  #删除行
            count = self.selection_Model.rowCount()  #获取行的数量
            self.selection_Model.insertRow(count)  #在末尾插入数据
            last_index = self.selection_Model.index(count, 0, QModelIndex()) #获取末尾的索引
            self.selection_Model.setData(last_index,string,Qt.DisplayRole) #设置末尾的数据
        self.view_clicked()  #控制按钮的激活与失效
    def btn_insert_clicked(self):  #插入按钮的槽函数
        while len(self.listView_1.selectedIndexes()):
            selectedIndexs_1 = self.listView_1.selectedIndexes()  # 获取选中的数据项的索引
            selectedIndex_2 = self.listView_2.selectedIndexes()  # 获取选中的数据项的索引
            index= selectedIndexs_1[0]
            string = self.reference_Model.data(index, Qt.DisplayRole)
            self.reference_Model.removeRow(index.row(), QModelIndex())
            row=selectedIndex_2[0].row()
            self.selection_Model.insertRow(row)
            index=self.selection_Model.index(row)
            self.selection_Model.setData(index, string, Qt.DisplayRole)
        self.view_clicked()
    def btn_delete_clicked(self):  #删除按钮的槽函数
        while len(self.listView_2.selectedIndexes()):
            selectedIndexes=self.listView_2.selectedIndexes()
            index= selectedIndexes[0]
            string = self.selection_Model.data(index, Qt.DisplayRole)
            self.selection_Model.removeRow(index.row(), QModelIndex())
            count = self.reference_Model.rowCount()
            self.reference_Model.insertRow(count)
            last_index = self.reference_Model.index(count, 0, QModelIndex())
            self.reference_Model.setData(last_index, string, Qt.DisplayRole)
        self.view_clicked()
        self.reference_Model.sort(0)  #排序
    def view_clicked(self):  #单击视图控件的槽函数，用于按钮的激活或失效
        n1=len(self.listView_1.selectedIndexes())  #获取选中数据项的数量
        n2 = len(self.listView_2.selectedIndexes())  #获取选中数据项的数量
        self.btn_add.setEnabled(n1)
        self.btn_insert.setEnabled(n1 and n2==1)
        self.btn_delete.setEnabled(n2)
if __name__ == '__main__':
    app=QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec())
