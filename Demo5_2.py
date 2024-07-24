import openpyxl   #Demo5_2.py
wbook = openpyxl.Workbook()  #创建工作簿实例对象
wsheet1 = wbook.active  #用wsheet1指向活动的工作表格
wsheet2 = wbook.create_sheet()  #创建工作表格对象wsheet2
wsheet3 = wbook.create_sheet("mySheet")  #创建工作表格对象，名称是mySheet
wsheet4 = wbook.create_sheet("mySheet1",0) #创建工作表格对象，名称是mySheet1，序号是0
wsheet5 = wbook.create_sheet("mySheet2",1) #创建工作表格对象，名称是mySheet2，序号是1
print(wsheet1.title,wsheet2.title,wsheet3.title,wsheet4.title,wsheet5.title)  #输出工作表格名称
print("活动工作表格的名称：",wbook.active.title)  #输出活动工作表格名称
print(wbook.sheetnames)  #输出工作表格名列表
wbook.save("d:\\myExcel.xlsx")  #存盘
#运行结果如下：
#Sheet Sheet1 mySheet mySheet1 mySheet2
#活动工作表格的名称： mySheet1
#['mySheet1', 'mySheet2', 'Sheet', 'Sheet1', 'mySheet']
